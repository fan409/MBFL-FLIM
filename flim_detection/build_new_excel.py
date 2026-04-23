import os
import shutil
import pandas as pd
from openpyxl import load_workbook
import math
import random
import hashlib

# =========================================================
# ⚙️ 配置运行模式
# =========================================================
STRATEGY = "mbert"        # "major" | "mbert"
RUN_MODE = "batch_all"  # "per_project" | "batch_all"
TARGET_PROJECT = "Chart"

# =========================================================
# ⚙️ FLIM Sus 修正策略
# =========================================================
# ✅ top-k% hard removal modes:
# - topk_llm_hard:            cand = (flim_ratio > 0), sort by flim_ratio desc, take top-k%
# - topk_random_hard:         cand = (flim_ratio > 0), random sample top-k% (reproducible)
# - topk_oracle_akf_hard:     cand = Oracle-FLIM from Excel (akf>0 && faulty_status==false), sort by akf desc, take top-k%
# - topk_oracle_random_hard:  cand = Oracle-FLIM from Excel, random sample top-k% (reproducible)

SUS_ADJUST_MODE = "topk_llm_hard"  # 这次只跑这个（你也可以换成 topk_oracle_random_hard）

HARD_THETA_MODE = "manual"  # "theta_star" | "manual"
HARD_THETA_MANUAL = 0
CLAMP_RATIO = True
USE_TOPK = SUS_ADJUST_MODE.startswith("topk")
# ✅ 一次跑完多个 top-k%
TOPK_LIST = [30,50,70]

THETA_FILE_NAME = "theta_star.txt"

# =========================================================
# ✅ 实验统计输出根目录（你指定的结构根路径）
# =========================================================
MITI_SAVE_ROOT = "/home/rs/WorkEx/Projects/SoftwareTesting/0ExperimentPlayGround/fxj/mitigate_top_k_flim"

# =========================================================
# ⚙️ 路径配置
# =========================================================
PATH_CONFIG = {
    "major": {
        "flim_eval_root": (
            "/home/rs/WorkEx/Projects/SoftwareTesting/MutationAnalysis/"
            "FLIMRecognitionResult/TraditionalMutation/major/"
            "Defects4J/D4JCleanCD4J/FLIM_Evaluation_added"
        ),
        "mbfl_input_root": (
            "/home/rs/WorkEx/Projects/SoftwareTesting/FaultLocalization/"
            "MBFL/Sus/Mutant/Defects4J/D4JCleanCD4J/"
            "TraditionalMutation/major/kill_type3"
        ),
        "mbfl_output_base": (
            "/home/rs/WorkEx/Projects/SoftwareTesting/FaultLocalization/"
            "MBFL/Sus/Mutant/Defects4J/D4JCleanCD4J/"
            "TraditionalMutation"
        ),
        "mbfl_output_prefix": "major_flim",
    },

    "mbert": {
        "flim_eval_root": (
            "/home/rs/WorkEx2/Projects/SoftwareTesting/MutationAnalysis/"
            "FLIMRecognitionResult/NeuralMutation/mBERT/"
            "Defects4J/D4JCleanCD4J/FLIM_Evaluation_added"
        ),
        "mbfl_input_root": (
            "/home/rs/WorkEx/Projects/SoftwareTesting/FaultLocalization/"
            "MBFL/Sus/Mutant/Defects4J/D4JCleanCD4J/"
            "NeuralMutation/mBERT/kill_type3"
        ),
        "mbfl_output_base": (
            "/home/rs/WorkEx/Projects/SoftwareTesting/FaultLocalization/"
            "MBFL/Sus/Mutant/Defects4J/D4JCleanCD4J/"
            "NeuralMutation"
        ),
        "mbfl_output_prefix": "mBERT_flim",
    }
}

FLIM_EVAL_ROOT  = PATH_CONFIG[STRATEGY]["flim_eval_root"]
MBFL_INPUT_ROOT = PATH_CONFIG[STRATEGY]["mbfl_input_root"]

# =========================================================
# ✅ 根据策略生成输出目录名 tag
# =========================================================
def mode_to_tag(mode: str) -> str:
    if mode == "topk_llm_hard":
        return "llm"
    if mode == "topk_random_hard":
        return "random"
    if mode == "topk_oracle_akf_hard":
        return "oracle_akf"
    if mode == "topk_oracle_random_hard":
        return "oracle_random"
    if mode == "soft":
        return "soft"
    if mode == "hard":
        return "hard"
    if mode == "none":
        return "none"
    if mode == "oracle":
        return "oracle_all"
    return mode

OUTPUT_TAG = mode_to_tag(SUS_ADJUST_MODE)

# =========================================================
# 工具函数
# =========================================================
def is_real_xlsx(fname):
    return (
        fname.endswith(".xlsx")
        and not fname.startswith("~$")
        and not fname.endswith(".xlsx.zip")
    )

def clamp01(x):
    try:
        v = float(x)
    except Exception:
        return 0.0
    return max(0.0, min(1.0, v))

def parse_bool_like(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(int(v))
    s = str(v).strip().lower()
    if s in ("true", "t", "yes", "y", "1"):
        return True
    if s in ("false", "f", "no", "n", "0", ""):
        return False
    return True

def stable_seed_int(seed_str: str) -> int:
    return int(hashlib.md5(seed_str.encode("utf-8")).hexdigest()[:8], 16)

# =========================================================
# ✅ LLM/Random Top-k removed_set 构造（version-level）
#    （不改你原逻辑；Oracle 改为 per-sheet topk）
# =========================================================
def build_topk_removed_set(mutant_df, project, version, mode, topk_pct):
    stats = {
        "total_mutants": 0,
        "cand_mutants": 0,
        "removed_mutants": 0,
        "removed_ratio_in_all": 0.0,
        "removed_ratio_in_cand": 0.0,
    }

    if mutant_df is None or len(mutant_df) == 0:
        return set(), stats

    v = str(version)
    vdf = mutant_df[mutant_df["version"] == v].copy()
    if len(vdf) == 0:
        return set(), stats

    vdf["mutant_id"] = vdf["mutant_id"].astype(str)
    stats["total_mutants"] = len(vdf)

    vdf["flim_ratio"] = pd.to_numeric(vdf["flim_ratio"], errors="coerce").fillna(0.0)
    if CLAMP_RATIO:
        vdf["flim_ratio"] = vdf["flim_ratio"].clip(0.0, 1.0)

    cand = vdf[vdf["flim_ratio"] > 0.0].copy()
    cand["Sus"] = pd.to_numeric(cand["Sus"], errors="coerce").fillna(0.0)
    stats["cand_mutants"] = len(cand)
    if len(cand) == 0:
        return set(), stats

    k = int(math.ceil(len(cand) * (float(topk_pct) / 100.0)))
    k = max(k, 1)

    if mode == "topk_llm_hard":
        cand = cand.sort_values(
            ["flim_ratio", "Sus", "mutant_id"],
            ascending=[False, True, True]
        )
        removed_set = set(cand.head(k)["mutant_id"].tolist())

    elif mode == "topk_random_hard":
        cand_ids = cand["mutant_id"].tolist()
        seed_str = f"{project}|{version}|{topk_pct}|{mode}"
        rng = random.Random(stable_seed_int(seed_str))

        if k >= len(cand_ids):
            removed_set = set(cand_ids)
        else:
            removed_set = set(rng.sample(cand_ids, k))
    else:
        return set(), stats

    stats["removed_mutants"] = len(removed_set)
    stats["removed_ratio_in_all"] = (stats["removed_mutants"] / stats["total_mutants"]) if stats["total_mutants"] > 0 else 0.0
    stats["removed_ratio_in_cand"] = (stats["removed_mutants"] / stats["cand_mutants"]) if stats["cand_mutants"] > 0 else 0.0
    return removed_set, stats

# =========================================================
# ✅ Sus 调整策略
# =========================================================
def adjust_sus(original_sus, flim_ratio, mode, theta=None, is_removed=False):
    try:
        s = float(original_sus)
    except Exception:
        s = 0.0

    p = clamp01(flim_ratio) if CLAMP_RATIO else float(flim_ratio)

    if mode == "none":
        return s
    if mode == "hard":
        if theta is None:
            raise ValueError("hard mode requires theta")
        return 0.0 if p > float(theta) else s
    if mode in ("topk_llm_hard", "topk_random_hard", "topk_oracle_akf_hard", "topk_oracle_random_hard"):
        return 0.0 if is_removed else s
    if mode == "soft":
        return s * (1.0 - p)
    if mode == "oracle":
        return s
    raise ValueError(f"Unknown SUS_ADJUST_MODE: {mode}")

# =========================================================
# ✅ process_copied_excel：按 sheet 统计（每个公式一个 sheet）
#    ✅ Oracle 两种策略：per-sheet topk，并且 oracle 更新 Sus 完全不依赖 flim_ratio_map
# =========================================================
def process_copied_excel(excel_path, version, flim_ratio_map, mode, theta=None, removed_set=None,
                        project=None, topk_pct=None):
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        print(f"[SKIP] {excel_path} (unreadable: {e})")
        return 0, 0, [], {
            "total_mutants": 0,
            "cand_mutants": 0,
            "removed_mutants": 0,
            "removed_ratio_in_all": 0.0,
            "removed_ratio_in_cand": 0.0,
        }

    modified_total = 0
    matched_total = 0
    sheet_stats = []

    # ✅ Oracle 总统计（跨 sheet 累加，仅用于输出统计，不参与 topk 选择）
    oracle_stats = {
        "total_mutants": 0,      # 这里定义为：所有参与 oracle cand 判断的行数总和（每个 sheet 每个 id 出现一次）
        "cand_mutants": 0,       # 严格满足 akf>0 && faulty_status==false 的行数总和
        "removed_mutants": 0,    # 在 cand 基础上按 topk_pct 选出的行数总和
        "removed_ratio_in_all": 0.0,
        "removed_ratio_in_cand": 0.0,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if headers is None:
            continue

        matched_sheet = 0
        modified_sheet = 0

        # 必须具备 mutant_id + Sus
        if "mutant_id" not in headers or "Sus" not in headers:
            continue

        mid_col = headers.index("mutant_id") + 1
        sus_col = headers.index("Sus") + 1

        # =========================================================
        # ✅ Oracle per-sheet topk removed_set_sheet 构造
        # cand = (akf>0 && faulty_status==false) 严格成立才进入 cand
        # 在 cand 基础上按 topk_pct 取 top-k%
        # =========================================================
        removed_set_sheet = set()
        sheet_total_rows = 0
        sheet_cand_list = []  # list of (mutant_id, akf_float)

        if mode in ("topk_oracle_akf_hard", "topk_oracle_random_hard"):
            # oracle 必须有 akf / faulty_status 两列
            if "akf" in headers and "faulty_status" in headers:
                akf_col = headers.index("akf") + 1
                fs_col  = headers.index("faulty_status") + 1

                for r in range(2, ws.max_row + 1):
                    mid = ws.cell(row=r, column=mid_col).value
                    if mid is None:
                        continue
                    mid = str(mid).strip()
                    if not mid:
                        continue

                    sheet_total_rows += 1

                    akf_val = ws.cell(row=r, column=akf_col).value
                    try:
                        akf_val = float(akf_val)
                    except Exception:
                        akf_val = 0.0

                    faulty_status_val = ws.cell(row=r, column=fs_col).value
                    fs = parse_bool_like(faulty_status_val)

                    # ✅ 严格 cand 范围：akf>0 且 faulty_status 为 false
                    if (akf_val > 0.0) and (fs == False):
                        sheet_cand_list.append((mid, akf_val))

                # cand 上取 top-k%
                if topk_pct is not None and len(sheet_cand_list) > 0:
                    k = int(math.ceil(len(sheet_cand_list) * (float(topk_pct) / 100.0)))
                    k = max(k, 1)

                    if mode == "topk_oracle_akf_hard":
                        sheet_cand_list.sort(key=lambda x: (-x[1], x[0]))  # akf desc, mutant_id asc
                        removed_set_sheet = set([mid for mid, _ in sheet_cand_list[:k]])
                    else:
                        cand_ids = [mid for mid, _ in sheet_cand_list]
                        # ✅ per-sheet 可复现实验：把 sheet_name 纳入种子
                        seed_str = f"{project}|{version}|{topk_pct}|{mode}|{sheet_name}"
                        rng = random.Random(stable_seed_int(seed_str))
                        removed_set_sheet = set(cand_ids) if k >= len(cand_ids) else set(rng.sample(cand_ids, k))

                # 累加 oracle 统计（跨 sheet）
                oracle_stats["total_mutants"] += sheet_total_rows
                oracle_stats["cand_mutants"] += len(sheet_cand_list)
                oracle_stats["removed_mutants"] += len(removed_set_sheet)

        # =========================================================
        # ✅ 逐行改 Sus
        # - Oracle：完全不依赖 flim_ratio_map，只依赖 removed_set_sheet
        # - 其他策略：保持你原来的 flim_ratio_map 匹配逻辑
        # =========================================================
        for r in range(2, ws.max_row + 1):
            mid = ws.cell(row=r, column=mid_col).value
            if mid is None:
                continue
            mid = str(mid).strip()
            if not mid:
                continue

            old_sus = ws.cell(row=r, column=sus_col).value

            # ✅ Oracle：真实值策略，不走 LLM flim_ratio_map 门槛
            if mode in ("topk_oracle_akf_hard", "topk_oracle_random_hard"):
                matched_sheet += 1
                p = 0.0  # 占位（oracle 不用 p）
                is_removed = (mid in removed_set_sheet)

            # ✅ 其他策略：沿用你原来的 LLM flim_ratio_map 匹配逻辑
            else:
                key = (str(version), str(mid))
                if key not in flim_ratio_map:
                    continue
                matched_sheet += 1
                p = flim_ratio_map[key]
                is_removed = (removed_set is not None and mid in removed_set)

            new_sus = adjust_sus(old_sus, p, mode=mode, theta=theta, is_removed=is_removed)

            try:
                old_val = float(old_sus)
            except Exception:
                old_val = None

            if old_val is None or abs(float(new_sus) - old_val) > 0.0:
                ws.cell(row=r, column=sus_col, value=float(new_sus))
                modified_sheet += 1

        matched_total += matched_sheet
        modified_total += modified_sheet

        sheet_stats.append({"formula": sheet_name, "matched": matched_sheet, "updated": modified_sheet})

    # oracle 总比例（仅用于输出统计）
    oracle_stats["removed_ratio_in_all"] = (
        oracle_stats["removed_mutants"] / oracle_stats["total_mutants"]
        if oracle_stats["total_mutants"] > 0 else 0.0
    )
    oracle_stats["removed_ratio_in_cand"] = (
        oracle_stats["removed_mutants"] / oracle_stats["cand_mutants"]
        if oracle_stats["cand_mutants"] > 0 else 0.0
    )

    wb.save(excel_path)
    return modified_total, matched_total, sheet_stats, oracle_stats

# =========================================================
# 获取项目列表
# =========================================================
if RUN_MODE == "per_project":
    projects = [TARGET_PROJECT]
else:
    projects = [
        d for d in os.listdir(FLIM_EVAL_ROOT)
        if os.path.isdir(os.path.join(FLIM_EVAL_ROOT, d))
        and os.path.exists(os.path.join(FLIM_EVAL_ROOT, d, THETA_FILE_NAME))
    ]

print(f"[STRATEGY] {STRATEGY}")
print(f"[MODE] SUS_ADJUST_MODE = {SUS_ADJUST_MODE}")
print(f"[TOPK_LIST] {TOPK_LIST}")
print(f"[PATH] FLIM_EVAL_ROOT  = {FLIM_EVAL_ROOT}")
print(f"[PATH] MBFL_INPUT_ROOT = {MBFL_INPUT_ROOT}")
print(f"[RUN_MODE] {RUN_MODE}, Projects = {len(projects)}")

# =========================================================
# 主流程
# =========================================================
for project in projects:
    print(f"\n[PROJECT] {project}")

    eval_dir = os.path.join(FLIM_EVAL_ROOT, project)
    mutant_csv = os.path.join(eval_dir, "mutant_aggregated.csv")

    mbfl_input = os.path.join(MBFL_INPUT_ROOT, project)

    # 1️⃣ 读取 mutant_aggregated.csv 构建 flim_ratio_map（仅供 LLM/soft/hard 等策略使用）
    flim_ratio_map = {}
    mutant_df = None

    if os.path.exists(mutant_csv):
        mutant_df = pd.read_csv(mutant_csv)
        mutant_df["project"] = mutant_df["project"].astype(str)
        mutant_df["version"] = mutant_df["version"].astype(str)
        mutant_df["mutant_id"] = mutant_df["mutant_id"].astype(str)

        mutant_df["flim_ratio"] = pd.to_numeric(mutant_df["flim_ratio"], errors="coerce").fillna(0.0)
        if CLAMP_RATIO:
            mutant_df["flim_ratio"] = mutant_df["flim_ratio"].clip(0.0, 1.0)

        flim_ratio_map = dict(zip(
            zip(mutant_df["version"], mutant_df["mutant_id"]),
            mutant_df["flim_ratio"]
        ))
    else:
        print(f"[WARN] mutant_aggregated.csv not found: {mutant_csv}")
        mutant_df = None
        flim_ratio_map = {}

    # 2️⃣ 依次跑 TOPK
    for TOPK_PCT in (TOPK_LIST if USE_TOPK else [None]):
        # ✅ 每个 topk 都是独立输出目录
        if USE_TOPK:
            MBFL_OUTPUT_ROOT = os.path.join(
                PATH_CONFIG[STRATEGY]["mbfl_output_base"],
                f"{PATH_CONFIG[STRATEGY]['mbfl_output_prefix']}_topk{TOPK_PCT}_{OUTPUT_TAG}",
                "kill_type3"
            )
        else:
            # ✅ soft / hard / none 走这里
            MBFL_OUTPUT_ROOT = os.path.join(
                PATH_CONFIG[STRATEGY]["mbfl_output_base"],
                f"{PATH_CONFIG[STRATEGY]['mbfl_output_prefix']}_{OUTPUT_TAG}",
                "kill_type3"
            )

        print(f"\n[TOPK] Running TOPK_PCT = {TOPK_PCT}%")
        print(f"[PATH] MBFL_OUTPUT_ROOT= {MBFL_OUTPUT_ROOT}")

        mbfl_output = os.path.join(MBFL_OUTPUT_ROOT, project)   # kill_type3/<project>
        os.makedirs(mbfl_output, exist_ok=True)

        summary_records = []
        formula_records_map = {}

        # 3️⃣ 逐 Excel：拷贝 → 修改（只动拷贝）
        for fname in sorted(os.listdir(mbfl_input)):
            if not is_real_xlsx(fname):
                continue

            src = os.path.join(mbfl_input, fname)
            dst = os.path.join(mbfl_output, fname)
            shutil.copy2(src, dst)

            version = fname.replace(f"{project}_", "").replace(".xlsx", "")

            # ✅ 非 Oracle：沿用你原来的 version-level removed_set（来自 flim_ratio>0）
            # ✅ Oracle：removed_set 不参与置0（置0由 per-sheet removed_set_sheet 完成）
            if SUS_ADJUST_MODE in ("topk_llm_hard", "topk_random_hard"):
                removed_set, topk_stats = build_topk_removed_set(
                    mutant_df, project, version, mode=SUS_ADJUST_MODE, topk_pct=TOPK_PCT
                )
            else:
                removed_set, topk_stats = set(), {
                    "total_mutants": 0,
                    "cand_mutants": 0,
                    "removed_mutants": 0,
                    "removed_ratio_in_all": 0.0,
                    "removed_ratio_in_cand": 0.0,
                }

            modified, matched, sheet_stats, oracle_stats = process_copied_excel(
                dst,
                version,
                flim_ratio_map,
                SUS_ADJUST_MODE,
                theta=None,
                removed_set=removed_set,
                project=project,
                topk_pct=TOPK_PCT
            )

            # ✅ Oracle 模式：统计用 oracle_stats（cand 严格=akf>0 && faulty_status==false，且在 cand 上取 topk%）
            if SUS_ADJUST_MODE in ("topk_oracle_akf_hard", "topk_oracle_random_hard"):
                stat_pack = oracle_stats
            else:
                stat_pack = topk_stats

            summary_records.append({
                "project": project,
                "version": version,
                "strategy": STRATEGY,
                "sus_adjust_mode": SUS_ADJUST_MODE,
                "topk_pct": TOPK_PCT,
                "total_mutants": stat_pack["total_mutants"],
                "cand_mutants": stat_pack["cand_mutants"],
                "removed_mutants": stat_pack["removed_mutants"],
                "removed_ratio_in_all": stat_pack["removed_ratio_in_all"],
                "removed_ratio_in_cand": stat_pack["removed_ratio_in_cand"],
                "matched_in_excel_total": matched,
                "sus_updated_total": modified,
                "output_excel": dst,
            })

            if sheet_stats:
                for ss in sheet_stats:
                    formula = ss["formula"]
                    row = {
                        "project": project,
                        "version": version,
                        "strategy": STRATEGY,
                        "sus_adjust_mode": SUS_ADJUST_MODE,
                        "topk_pct": TOPK_PCT,
                        "formula": formula,
                        "matched_in_excel_sheet": ss["matched"],
                        "sus_updated_sheet": ss["updated"],
                        "output_excel": dst,
                        "total_mutants": stat_pack["total_mutants"],
                        "cand_mutants": stat_pack["cand_mutants"],
                        "removed_mutants": stat_pack["removed_mutants"],
                        "removed_ratio_in_all": stat_pack["removed_ratio_in_all"],
                        "removed_ratio_in_cand": stat_pack["removed_ratio_in_cand"],
                    }
                    formula_records_map.setdefault(formula, []).append(row)

            extra_info = (
                f"TOPK={TOPK_PCT}%,"
                f"Cand={stat_pack['cand_mutants']}/{stat_pack['total_mutants']},"
                f"Removed={stat_pack['removed_mutants']},"
                f"Rall={stat_pack['removed_ratio_in_all']:.3f},"
                f"Rcand={stat_pack['removed_ratio_in_cand']:.3f}"
            )

            mutants_in_agg = stat_pack["total_mutants"]

            print(
                f"[OK] {fname}: "
                f"MutantsInAgg={mutants_in_agg}, "
                f"MatchedInExcel={matched}, "
                f"SusUpdated={modified}, "
                f"{extra_info}"
            )

        # 4️⃣ 存统计：按你要求的目录结构
        # /.../mitigate_top_k_flim/<strategy>/<mode>/topk_<pct>/<project>/
        if USE_TOPK:
            out_dir = os.path.join(
                MITI_SAVE_ROOT,
                STRATEGY,
                SUS_ADJUST_MODE,
                f"topk_{TOPK_PCT}",
                project
            )
        else:
            out_dir = os.path.join(
                MITI_SAVE_ROOT,
                STRATEGY,
                SUS_ADJUST_MODE,
                project
            )
        os.makedirs(out_dir, exist_ok=True)

        if summary_records:
            out_csv = os.path.join(out_dir, "mitigation_stats.csv")
            pd.DataFrame(summary_records).to_csv(out_csv, index=False)
            print(f"[SAVE] mitigation stats -> {out_csv}")

        if formula_records_map:
            by_dir = os.path.join(out_dir, "by_formula")
            os.makedirs(by_dir, exist_ok=True)

            for formula_name, rows in formula_records_map.items():
                safe_name = str(formula_name).strip().replace("/", "_").replace("\\", "_").replace(" ", "_")
                out_path = os.path.join(by_dir, f"{safe_name}.csv")
                pd.DataFrame(rows).to_csv(out_path, index=False)

            print(f"[SAVE] by_formula csvs -> {by_dir}")

        meta_path = os.path.join(out_dir, "meta.txt")
        with open(meta_path, "w") as f:
            f.write(f"STRATEGY={STRATEGY}\n")
            f.write(f"RUN_MODE={RUN_MODE}\n")
            f.write(f"TARGET_PROJECT={TARGET_PROJECT}\n")
            f.write(f"SUS_ADJUST_MODE={SUS_ADJUST_MODE}\n")
            f.write(f"TOPK_PCT={TOPK_PCT}\n")
            f.write(f"FLIM_EVAL_ROOT={FLIM_EVAL_ROOT}\n")
            f.write(f"MBFL_INPUT_ROOT={MBFL_INPUT_ROOT}\n")
            f.write(f"MBFL_OUTPUT_ROOT={MBFL_OUTPUT_ROOT}\n")
            f.write(f"OUTPUT_TAG={OUTPUT_TAG}\n")

        print(f"[SAVE] meta -> {meta_path}")

print("\n[DONE] All projects processed.")
